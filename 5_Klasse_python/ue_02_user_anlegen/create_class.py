import argparse
import random
import sys

from openpyxl import load_workbook
import os.path

from logging.handlers import RotatingFileHandler
import logging

logger = logging.getLogger(__name__)
handler = logging.handlers.RotatingFileHandler("create_class.log", maxBytes=10000, backupCount=5)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(formatter)

logger.addHandler(handler)
logger.addHandler(stream_handler)

""" Beispiel Excel auslesen 
wb = load_workbook(xlsfilename, read_only=True)
ws = wb[wb.sheetnames[0]]
for row in ws.iter_rows(min_row=7):
x = row[3] # Element in der Zeile lesen -- Objekt
y = x.value # Wert in der Zelle
z = x.coordinate # "Name" der Zelle
ws['A1'] = 42 # Schreiben mit Zellenname
"""

user_to_password = []
verbosity = False


def read_class_excel(filename: str) -> str:
    """
        reads an excel file and gives back the values of each line

        :param filename: file to read
        :return: lines of the excel file as values
        """
    lines = []
    wb = load_workbook(filename, read_only=True)
    ws = wb[wb.sheetnames[0]] # Hier kann man probably andere Zahlen angeben für andere Sheets in der gleichen Datei
    for row in ws.iter_rows(min_row=2):
        if row[0].value == None:
            break
        lines.append((row[0].value, row[1].value, row[2].value))
    return lines


def create_files(excel_file: str) -> None:
    """
    creates and writes the required Files

    :return: The required files at the place where the program is started
    """
    global user_to_password
    global verbosity
    global logger
    lines = read_class_excel(excel_file)
    class_user_script = create_class_users_script(lines)  # quasi ein String den man dann einfach in ein File schreibt
    class_delete_script = create_class_delete_script()
    with open('create_class_script.sh', 'w') as f1:
        f1.write(class_user_script)
        logger.info("class_user_script created")
        if verbosity:
            print('class_user_script created')
    with open('delete_class_script.sh', 'w') as f1:
        f1.write(class_delete_script)
        logger.info("class_delete_script created")
        if verbosity:
            print("class_delete_script created")
    with open('user_password_list.txt', 'w') as f1:
        f1.write(get_user_to_passwd_list())
        logger.info("user_password_list created")
        if verbosity:
            print("user_password_list created")
    # print(line, 'ressources/class_user_script.txt')


def create_class_users_script(lines: str) -> str:
    """
        creates the actual bash script to build the required users

        :param lines: lines of the excel file that need to be written as a script
        :return: returns the script as a string so that it can be written into a file
        """
    global user_to_password
    global verbosity
    global logger
    script = ''
    # zuerst Variablen für alle Werte die man so braucht
    # dann eine Zeile die in ein File schreibt, und wirklich die script Zeilen erstellt
    systemgroups = 'cdrom,plugdev,sambashare'
    home_shell = '/bin/bash'
    script += 'groupadd lehrer\n'
    script += 'groupadd seminar\n'
    script += 'groupadd klasse\n'
    script += f'useradd -d /home/lehrer -c \"Lehrer\" -m -g lehrer -G {systemgroups} -s {home_shell} lehrer\n'
    script += f'useradd -d /home/seminar -c \"Seminar\" -m -g seminar -G {systemgroups} -s {home_shell} seminar\n'
    script += 'mkdir /home/klassen\n'
    random_chars = ['!', '%', '\(', '\)', ',', '.', '_', '-', '=', '^', '#']
    for line in lines:
        class_name = str(line[0]).lower()
        room_number = (str(line[1]).lower())[:3]
        class_teacher = str(line[2]).lower()
        username = 'k' + class_name
        username = username.replace('ä', 'ae')
        username = username.replace('ö', 'oe')
        username = username.replace('ü', 'ue')
        username = username.replace('ß', 'ss')
        gecos_field = class_name + '_' + class_teacher
        home_directory = '/home/klassen/' + username
        password = class_name + random.choice(random_chars) + room_number + random.choice(
            random_chars) + class_teacher + random.choice(random_chars)
        user_to_password.append((username, password))
        # script += f'groupadd {username}\n'
        script += f'useradd -d {home_directory} -c \"{gecos_field}\" -m -g {username} -G {systemgroups} -s {home_shell} {username}\n'
        script += f'echo {username}:{password} | chpasswd\n'
        logger.info(f'User {username} with password created!')
        if verbosity:
            script += f'echo User {username} with password created!\n'
            print('echo User {username} with password created!')
    return script


def create_class_delete_script() -> str:
    """
        creates the delete script matching the create script

        :return: the delete script as a string to be writeable
        """
    global user_to_password
    global verbosity
    global logger
    script = ''
    for user in user_to_password:
        username = user[0]
        script += f'userdel -r {username}\n'
        logger.info(f'User {username} with password deleted!')
        if verbosity:
            script += f'echo User {username} with password deleted!\n'
            print('User {username} with password deleted!')
    return script


def get_user_to_passwd_list() -> str:
    """
        writes a list for every user with its password

        :return: the user-password list as a string
        """
    global user_to_password
    global verbosity
    global logger
    list = ''
    for user, password in user_to_password:
        list += str(user) + ';' + str(password) + '\n'
        logger.info(f'User {user} and its password in user_to_password list')
        if verbosity:
            print(f'User {user} and its password in user_to_password list')
    return list


# create_files('ressources/Klassenraeume_2023.xlsx')


def main():
    """
        arg-parser for the program to be runnable

        :return: -
        """
    global verbosity
    global logger
    parser = argparse.ArgumentParser()
    parser.add_argument("filename", type=str, help="filename")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("-v", "--verbose", help="add log verbosity", action="store_true")
    group.add_argument("-q", "--quiet", help="turn off log verbosity", action="store_true")

    args = parser.parse_args()
    logger.setLevel(logging.INFO)
    if args.verbose:
        # global verbosity
        verbosity = True
        print("Verbosity turned on")

    if args.quiet:
        # global verbosity
        verbosity = False
        print('Verbosity turned off')

    if args.filename:
        create_files(args.filename)


if __name__ == "__main__":
    main()
