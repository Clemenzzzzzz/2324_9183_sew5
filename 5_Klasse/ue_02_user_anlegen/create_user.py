import argparse
import random
import re
import sys
from collections import Counter

import unicodedata
import unidecode
from openpyxl import load_workbook
import os.path

from logging.handlers import RotatingFileHandler
import logging

logger = logging.getLogger(__name__)
handler = logging.handlers.RotatingFileHandler("create_class.log", maxBytes=10000, backupCount=5)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

stream_handler = logging.StreamHandler(sys.stdout)
stream_handler.setFormatter(formatter)

logger.addHandler(handler)
logger.addHandler(stream_handler)

""" Beispiel Excel auslesen 
wb = load_workbook(xlsfilename, read_only=True)
ws = wb[wb.sheetnames[0]]
for row in ws.iter_rows(min_row=7):
x = row[3] # Element in der Zeile lesen -- Objekt
y = x.value # Wert in der Zelle
z = x.coordinate # "Name" der Zelle
ws['A1'] = 42 # Schreiben mit Zellenname
"""

user_to_password = []
existing_user_names = []
verbosity = False


def read_name_excel(filename):
    """
    reads an excel file and gives back the values of each line

    :param filename: file to read
    :return: lines of the excel file as values
    """
    lines = []
    wb = load_workbook(filename, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == None:
            break
        # z = x.coordinate  # "Name" der Zelle
        lines.append((row[0].value, row[1].value, row[2].value, row[3].value))
    # ws['A1'] = 42  # Schreiben mit Zellenname
    return lines


def create_files(excel_file):
    """
    creates and writes the required Files

    :return: The required files at the place where the program is started
    """
    global user_to_password
    global verbosity
    global logger
    lines = read_name_excel(excel_file)
    real_user_script = create_real_users_script(lines)  # quasi ein String den man dann einfach in ein File schreibt
    real_delete_script = create_real_delete_script()
    with open('create_user_script.sh', 'w', encoding='utf-8') as f1:
        print(real_user_script)
        f1.write(real_user_script)
        logger.info("real_user_script created")
        if verbosity:
            print('real_user_script created')
    with open('delete_user_script.sh', 'w', encoding='utf-8') as f1:
        f1.write(real_delete_script)
        logger.info("real_delete_script created")
        if verbosity:
            print("real_delete_script created")
    with open('user_password_list.txt', 'w', encoding='utf-8') as f1:
        f1.write(get_user_to_passwd_list())
        logger.info("user_password_list created")
        if verbosity:
            print("user_password_list created")
    # print(line, 'ressources/class_user_script.txt')


def create_real_users_script(lines):
    """
    creates the actual bash script to build the required users

    :param lines: lines of the excel file that need to be written as a script
    :return: returns the script as a string so that it can be written into a file
    """
    global user_to_password
    global verbosity
    global logger
    script = ''
    # zuerst Variablen für alle Werte die man so braucht
    # dann eine Zeile die in ein File schreibt, und wirklich die script Zeilen erstellt
    systemgroups = 'cdrom,plugdev,sambashare'
    home_shell = '/bin/bash'
    script += 'mkdir /home\n'
    random_chars = ['!', '%', '\(', '\)', ',', '.', '_', '-', '=', '^', '#']
    for line in lines:
        print('hier is die line')
        print(line)
        firstname = str(line[0])
        lastname = str(line[1])
        group = str(line[2])
        user_class = str(line[3])
        username = get_valid_username(lastname)
        username = handle_double_users(username)
        gecos_field = firstname + '_' + lastname
        home_directory = '/home/' + username
        password = username + random.choice(random_chars) + group + random.choice(
            random_chars) + user_class + random.choice(random_chars)
        user_to_password.append((username, password))
        # script += f'groupadd {username}\n'
        script += f'useradd -d {home_directory} -c \"{gecos_field}\" -m -g {username} -G {systemgroups},{group} -s {home_shell} {username}\n'
        script += f'echo {username}:{password} | chpasswd\n'
        logger.info(f'User {username} with password created!')
        if verbosity:
            script += f'echo User {username} with password created!\n'
            print('echo User {username} with password created!')
    return script


def handle_double_users(username):
    """
    method to check if a username is already there

    :param username: username that needs to be checkes
    :return: the new username if a username is double
    >>> user1 = handle_double_users('august')
    >>> user2 = handle_double_users('clemens')
    >>> user3 = handle_double_users('august')
    >>> user4 = handle_double_users('august')
    >>> print(user1)
    august
    >>> print(user2)
    clemens
    >>> print(user3)
    august1
    >>> print(user4)
    august2
    """
    if username in existing_user_names:
        index = 1
        while f'{username}{index}' in existing_user_names:
            index += 1
        username = username + str(index)
        existing_user_names.append(username)
    else:
        existing_user_names.append(username)
    return username


def get_valid_username(lastname):
    """
    changes a lastname to a valid username in the script

    :param lastname: the lastname as given in the excel file
    :return: a valid username as string
    >>> user1 = get_valid_username('Hodina')
    >>> user2 = get_valid_username('Hörandl')
    >>> user3 = get_valid_username('IRÉNÉE')
    >>> user4 = get_valid_username('Núñez Gómez')
    >>> print(user1)
    hodina
    >>> print(user2)
    hoerandl
    >>> print(user3)
    irenee
    >>> print(user4)
    nunez_gomez
    """
    username = lastname.lower()
    # username = unidecode.unidecode(username)
    username = username.replace(' ', '_')
    username = username.replace('ä', 'ae')
    username = username.replace('ö', 'oe')
    username = username.replace('ü', 'ue')
    username = username.replace('ß', 'ss')
    username = shave_marks(username)
    username = re.sub(r'[^a-z0-9_]', '', username)
    return username


def shave_marks(txt):
    """Remove all diacritic marks"""
    norm_txt = unicodedata.normalize('NFD', txt)
    shaved = ''.join(c for c in norm_txt if not unicodedata.combining(c))
    return unicodedata.normalize('NFC', shaved)


# TODO tests siehe angbabe

def create_real_delete_script():
    """
    creates the delete script matching the create script

    :return: the delete script as a string to be writeable
    """
    global user_to_password
    global verbosity
    global logger
    script = ''
    for user in user_to_password:
        username = user[0]
        script += f'userdel -r {username}\n'
        logger.info(f'User {username} with password deleted!')
        if verbosity:
            script += f'echo User {username} with password deleted!\n'
            print('User {username} with password deleted!')
    return script


def get_user_to_passwd_list():  # TODO das auch als excel
    """
    writes a list for every user with its password

    :return: the user-password list as a string
    """
    global user_to_password
    global verbosity
    global logger
    list = ''
    for user, password in user_to_password:
        list += str(user) + ';' + str(password) + '\n'
        logger.info(f'User {user} and its password in user_to_password list')
        if verbosity:
            print(f'User {user} and its password in user_to_password list')
    return list


# create_files('ressources/Klassenraeume_2023.xlsx')


def main():
    """
    arg-parser for the program to be runnable

    :return: -
    """
    global verbosity
    global logger
    parser = argparse.ArgumentParser()
    parser.add_argument("filename", type=str, help="filename")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("-v", "--verbose", help="add log verbosity", action="store_true")
    group.add_argument("-q", "--quiet", help="turn off log verbosity", action="store_true")

    args = parser.parse_args()
    logger.setLevel(logging.INFO)
    if args.verbose:
        # global verbosity
        verbosity = True
        print("Verbosity truned on")

    if args.quiet:
        # global verbosity
        verbosity = False
        print('Verbosity turned off')

    if args.filename:
        create_files(args.filename)


if __name__ == "__main__":
    main()
